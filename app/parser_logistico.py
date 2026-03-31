"""
app/parser_logistico.py

Lê uma planilha logística e cria os registros EmpreendimentoLogistico.

Estrutura da planilha:
  - Linha com Nº preenchido = empreendimento principal
  - Linhas seguintes com Nº vazio = tipologias de módulos do mesmo empreendimento

Regras de agrupamento:
  - módulos / ocupados / disponíveis / vacância → vêm da linha RAIZ (representam o todo)
  - abl_m2 → SOMA da linha raiz + todas as tipologias
  - Tipologias NÃO são salvas como registros separados no banco

Geocoding:
  - Busca pelo NOME do empreendimento + cidade + PE (não pelo endereço da planilha)
  - Ex: "Armazenna 1, Jaboatão dos Guararapes, PE"
  - Isso dá resultados muito mais precisos que a BR-101
"""

import math
import logging
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from app.models import Dashboard, EmpreendimentoLogistico
from app.utils import buscar_coordenadas

logger = logging.getLogger(__name__)

# Estado fixo — todas as cidades desta planilha são PE
ESTADO = 'PE'

# ── Colunas esperadas (índice 0-based) ──────────────────────────────────────
COL = {
    'numero':            0,
    'nome':              1,
    'endereco':          2,
    'cidade':            3,
    'construtora':       4,
    'proprietario':      5,
    'fase_obra':         6,
    'num_galpoes':       7,
    'num_modulos':       8,
    'modulos_ocupados':  9,
    'modulos_disponiveis': 10,
    'preco_m2_locacao':    11,
    'preco_m2_condominio': 12,
    'preco_m2_iptu':       13,
    'preco_locacao':       14,
    'preco_condominio':    15,
    'preco_iptu':          16,
    'area_modulo_m2':      17,
    'abl_m2':              18,
    'vacancia':            19,
}


# ── Helpers de conversão ────────────────────────────────────────────────────

def _safe_decimal(val):
    if val is None or str(val).strip().upper() in ('NI', '', 'N/A', '-'):
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _safe_float(val):
    if val is None or str(val).strip().upper() in ('NI', '', 'N/A', '-'):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val):
    if val is None or str(val).strip().upper() in ('NI', '', 'N/A', '-'):
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi    = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ── Parser principal ─────────────────────────────────────────────────────────

def importar_planilha_logistica(arquivo, dashboard: Dashboard) -> dict:
    """
    Lê o xlsx e importa para o banco — 1 registro por empreendimento.

    Geocoding: busca pelo NOME + cidade + PE (não pelo endereço bruto da planilha).

    Retorna:
        empreendimentos_criados : int
        erros                   : list[str]
    """
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(min_row=2, values_only=True))  # pula cabeçalho

    # ── Passo 1: agrupar linhas por empreendimento ──────────────────────────
    grupos = []
    atual  = None

    for idx, row in enumerate(linhas, start=2):
        numero = _safe_int(row[COL['numero']])
        nome   = str(row[COL['nome']] or '').strip()

        if not nome or nome.upper() in ('NI', 'NAN'):
            continue

        if numero is not None:
            # Nova raiz — salva a anterior
            if atual is not None:
                grupos.append(atual)

            abl_raiz = _safe_float(row[COL['abl_m2']]) or 0

            atual = {
                'linha_idx':   idx,
                'numero':      numero,
                'nome':        nome,
                'endereco':    str(row[COL['endereco']] or '').strip() or None,
                'cidade':      str(row[COL['cidade']]   or '').strip() or None,
                'construtora': str(row[COL['construtora']] or '').strip() or None,
                'proprietario': str(row[COL['proprietario']] or '').strip() or None,
                'fase_obra':   str(row[COL['fase_obra']] or '').strip() or None,
                'num_galpoes':         _safe_int(row[COL['num_galpoes']]),
                'num_modulos':         _safe_int(row[COL['num_modulos']]),
                'modulos_ocupados':    _safe_int(row[COL['modulos_ocupados']]),
                'modulos_disponiveis': _safe_int(row[COL['modulos_disponiveis']]),
                'vacancia':            _safe_float(row[COL['vacancia']]),
                'preco_m2_locacao':    _safe_decimal(row[COL['preco_m2_locacao']]),
                'preco_m2_condominio': _safe_decimal(row[COL['preco_m2_condominio']]),
                'preco_m2_iptu':       _safe_decimal(row[COL['preco_m2_iptu']]),
                'preco_locacao':       _safe_decimal(row[COL['preco_locacao']]),
                'preco_condominio':    _safe_decimal(row[COL['preco_condominio']]),
                'preco_iptu':          _safe_decimal(row[COL['preco_iptu']]),
                'area_modulo_m2':      _safe_float(row[COL['area_modulo_m2']]),
                'abl_m2': abl_raiz,
            }

        else:
            # Tipologia — acumula ABL no pai
            if atual is not None:
                atual['abl_m2'] += _safe_float(row[COL['abl_m2']]) or 0

    if atual is not None:
        grupos.append(atual)

    # ── Passo 2: calcular vacância quando vier None ─────────────────────────
    for g in grupos:
        if g['vacancia'] is None:
            mod  = g['num_modulos']
            disp = g['modulos_disponiveis']
            if mod and disp is not None:
                g['vacancia'] = round(disp / mod, 6)

    # ── Passo 3: geocoding pelo NOME do empreendimento ──────────────────────
    #
    # Em vez de buscar pelo endereço bruto da planilha (ex: "BR 101 Sul"),
    # buscamos pelo NOME + cidade + estado, que funciona como uma pesquisa
    # no Google Maps. Ex: "Armazenna 1, Jaboatão dos Guararapes, PE"
    #
    ref_lat = dashboard.ref_latitude
    ref_lon = dashboard.ref_longitude
    tem_ref = ref_lat is not None and ref_lon is not None

    for g in grupos:
        lat, lon = None, None
        cidade = g['cidade'] or ''
        nome   = g['nome']

        # Monta a query como: "Nome do Empreendimento, Cidade, PE"
        query_nome   = nome
        query_cidade = cidade

        try:
            # buscar_coordenadas(rua, bairro, cidade)
            # Passamos nome como "rua" e cidade+estado como "cidade"
            # para que a função monte: "Nome, , Cidade PE"
            cidade_com_estado = f"{cidade}, {ESTADO}" if cidade else ESTADO
            lat, lon = buscar_coordenadas(query_nome, '', cidade_com_estado)

            if lat and lon:
                logger.info(f"✅ Geocoding OK: {nome} → {lat:.5f}, {lon:.5f}")
            else:
                # Fallback: tenta só com a cidade
                logger.warning(f"⚠️  Geocoding sem resultado para '{nome}', tentando só cidade...")
                lat, lon = buscar_coordenadas(cidade_com_estado, '', cidade_com_estado)

        except Exception as e:
            logger.warning(f"Geocoding falhou para '{nome}': {e}")

        g['latitude']  = lat
        g['longitude'] = lon

        distancia_km = None
        if tem_ref and lat and lon:
            distancia_km = round(_haversine(lat, lon, ref_lat, ref_lon), 2)
        g['distancia_ref_km'] = distancia_km

    # ── Passo 4: salvar no banco ─────────────────────────────────────────────
    empreendimentos_criados = 0
    erros = []

    for g in grupos:
        try:
            EmpreendimentoLogistico.objects.create(
                dashboard=dashboard,
                empreendimento_principal=None,
                numero=g['numero'],
                nome=g['nome'],
                endereco=g['endereco'],
                cidade=g['cidade'],
                construtora=g['construtora'],
                proprietario=g['proprietario'],
                fase_obra=g['fase_obra'],
                num_galpoes=g['num_galpoes'],
                num_modulos=g['num_modulos'],
                modulos_ocupados=g['modulos_ocupados'],
                modulos_disponiveis=g['modulos_disponiveis'],
                preco_m2_locacao=g['preco_m2_locacao'],
                preco_m2_condominio=g['preco_m2_condominio'],
                preco_m2_iptu=g['preco_m2_iptu'],
                preco_locacao=g['preco_locacao'],
                preco_condominio=g['preco_condominio'],
                preco_iptu=g['preco_iptu'],
                area_modulo_m2=g['area_modulo_m2'],
                abl_m2=g['abl_m2'],
                vacancia=g['vacancia'],
                latitude=g['latitude'],
                longitude=g['longitude'],
                distancia_ref_km=g['distancia_ref_km'],
            )
            empreendimentos_criados += 1
            logger.info(f"Criado: {g['nome']} | ABL={g['abl_m2']:.0f} m² | lat={g['latitude']} lon={g['longitude']}")

        except Exception as e:
            erros.append(f"Linha {g['linha_idx']}: erro ao salvar '{g['nome']}' — {e}")
            logger.exception(f"Erro ao salvar '{g['nome']}'")

    return {
        'empreendimentos_criados': empreendimentos_criados,
        'tipologias_criadas': 0,
        'erros': erros,
    }